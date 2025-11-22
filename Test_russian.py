# -*- coding: utf-8 -*-
import os
import sys
import openpyxl
import requests
import pygame
import tkinter as tk
from tkinter import ttk, messagebox
from urllib.parse import quote
import threading

def resource_path(relative_path):
    """Get path relative to exe or script folder"""
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

class RussianVocabularyApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Russian Vocabulary")
        self.root.geometry("865x600")
        
        # Initialize variables
        self.use_proxy = tk.BooleanVar(value=True)
        self.proxy_url = tk.StringVar(value="http://127.0.0.1:7890")
        self.status_var = tk.StringVar(value="Ready")
        self.current_selection = 0
        self.current_image = None
        self.level_var = tk.StringVar(value="A1")  # Default level
        self.current_columns = 6  # Default columns
        
        # Level data
        self.levels = ["A1", "A2", "B1", "B2", "C1", "C2"]
        self.level_words = {
            "A1": [
                ("привет", "Hello"), ("пока", "Goodbye"), ("спасибо", "Thank you"),
                ("пожалуйста", "Please"), ("да", "Yes"), ("нет", "No"),
                ("хорошо", "Good"), ("плохо", "Bad"), ("извините", "Excuse me"),
                ("я", "I"), ("ты", "You"), ("он", "He"), ("она", "She")
            ],
            "A2": [
                ("говорить", "to speak"), ("читать", "to read"), ("писать", "to write"),
                ("слушать", "to listen"), ("понимать", "to understand"), ("знать", "to know"),
                ("думать", "to think"), ("работать", "to work"), ("учиться", "to study")
            ],
            "B1": [
                ("обсуждать", "to discuss"), ("предлагать", "to suggest"), ("соглашаться", "to agree"),
                ("отказываться", "to refuse"), ("объяснять", "to explain"), ("доказывать", "to prove")
            ],
            "B2": [
                ("анализировать", "to analyze"), ("оценивать", "to evaluate"), ("аргументировать", "to argue"),
                ("критиковать", "to criticize"), ("интерпретировать", "to interpret")
            ],
            "C1": [
                ("абстрагироваться", "to abstract"), ("концептуализировать", "to conceptualize"),
                ("систематизировать", "to systematize"), ("оптимизировать", "to optimize")
            ],
            "C2": [
                ("идентифицировать", "to identify"), ("квалифицировать", "to qualify"),
                ("квантовать", "to quantize"), ("трансформировать", "to transform")
            ]
        }
        
        # Initialize pygame mixer
        pygame.mixer.init()
        
        # Load data
        self.setup_directories()
        self.load_level_data(self.level_var.get())
        
        # Create UI
        self.create_widgets()
        
        # Bind keyboard events
        self.bind_keyboard_events()
        
        # Refresh word list to show columns on startup
        self.refresh_word_list()
        
    def setup_directories(self):
        """Setup directories for all levels"""
        self.base_dir = resource_path("")
        self.words_dir = resource_path("words")
        os.makedirs(self.words_dir, exist_ok=True)
        
        # Create subdirectories for each level
        for level in self.levels:
            level_dir = os.path.join(self.words_dir, f"word_{level}")
            os.makedirs(level_dir, exist_ok=True)
            
            # Create audio subdirectory for each level
            audio_dir = os.path.join(level_dir, "audio_files")
            os.makedirs(audio_dir, exist_ok=True)
            
            # Create Excel file if it doesn't exist
            excel_path = os.path.join(level_dir, f"word_{level}.xlsx")
            if not os.path.exists(excel_path):
                self.create_level_excel(level, excel_path)
        
    def create_level_excel(self, level, path):
        """Create Excel file for specific level"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Words"

        ws.cell(row=1, column=1, value="Russian")
        ws.cell(row=1, column=2, value="English")

        # Add words for this level
        words = self.level_words.get(level, [])
        for i, (ru, en) in enumerate(words, start=2):
            ws.cell(row=i, column=1, value=ru)
            ws.cell(row=i, column=2, value=en)

        wb.save(path)
        print(f"Created Excel file for level {level}: {path}")
        
    def load_level_data(self, level):
        """Load data for specific level"""
        level_dir = os.path.join(self.words_dir, f"word_{level}")
        excel_path = os.path.join(level_dir, f"word_{level}.xlsx")
        
        if not os.path.exists(excel_path):
            self.create_level_excel(level, excel_path)
            
        self.wb = openpyxl.load_workbook(excel_path)
        self.sheet = self.wb.active
        
        # Load words into memory
        self.words = []
        for row in self.sheet.iter_rows(min_row=2):
            russian = row[0].value
            english = row[1].value
            if russian and english:
                self.words.append((russian, english))
                
        print(f"Loaded {len(self.words)} words for level {level}")
        self.current_selection = 0  # Reset selection when level changes
        
    def get_proxy(self):
        """Get proxy configuration if enabled"""
        if self.use_proxy.get() and self.proxy_url.get().strip():
            return self.proxy_url.get().strip()
        return None
        
    def get_audio_dir(self):
        """Get audio directory for current level"""
        level = self.level_var.get()
        level_dir = os.path.join(self.words_dir, f"word_{level}")
        return os.path.join(level_dir, "audio_files")
        
    def download_audio(self, text, language):
        """Download audio file"""
        safe_text = "".join(c if c.isalnum() else "_" for c in text)
        audio_dir = self.get_audio_dir()
        filename = os.path.join(audio_dir, f"{language}_{safe_text}.mp3")
        
        if os.path.exists(filename):
            return filename
            
        try:
            encoded_text = quote(text)
            if language == "ru":
                url = f"https://translate.google.com/translate_tts?ie=UTF-8&tl=ru&client=tw-ob&q={encoded_text}"
            else:
                url = f"https://translate.google.com/translate_tts?ie=UTF-8&tl=en&client=tw-ob&q={encoded_text}"
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                'Referer': 'https://translate.google.com/'
            }
            
            # Proxy configuration
            proxy = self.get_proxy()
            proxies = None
            if proxy:
                proxies = {
                    'http': proxy,
                    'https': proxy
                }
            
            self.update_status(f"Downloading {language} audio: {text}")
            response = requests.get(url, headers=headers, timeout=30, proxies=proxies)
            
            if response.status_code == 200:
                with open(filename, 'wb') as f:
                    f.write(response.content)
                return filename
            else:
                self.update_status(f"Download failed: {text} (Status: {response.status_code})")
                
        except Exception as e:
            self.update_status(f"Download failed: {e}")
        
        return None
        
    def play_audio(self, filename):
        """Play audio file"""
        try:
            pygame.mixer.music.load(filename)
            pygame.mixer.music.play()
        except Exception as e:
            messagebox.showerror("Error", f"Play audio failed: {e}")
            
    def update_status(self, message):
        """Update status message"""
        self.status_var.set(message)
        self.root.update_idletasks()
        print(message)
        
    def on_download_all(self):
        """Download all audio files for current level"""
        def download_thread():
            self.download_btn.config(state="disabled")
            success_count = 0
            total_count = len(self.words)
            
            for i, (russian, english) in enumerate(self.words):
                self.update_status(f"Downloading {i+1}/{len(self.words)}: {russian}")
                
                # Download Russian audio
                if self.download_audio(russian, "ru"):
                    success_count += 1
                        
            self.update_status(f"Download completed: {success_count}/{total_count} files")
            self.download_btn.config(state="normal")
            messagebox.showinfo("Download Complete", f"Downloaded {success_count}/{total_count} audio files for level {self.level_var.get()}")
            
        thread = threading.Thread(target=download_thread)
        thread.daemon = True
        thread.start()
        
    def play_current_word(self):
        """Play the currently selected Russian word"""
        if 0 <= self.current_selection < len(self.words):
            russian_word = self.words[self.current_selection][0]
            filename = self.download_audio(russian_word, "ru")
            if filename:
                self.play_audio(filename)
                self.update_status(f"Playing: {russian_word}")
            else:
                messagebox.showerror("Error", f"Cannot play audio for: {russian_word}")
        
    def update_selection(self):
        """Update the visual selection and English translation"""
        # Update Russian word buttons color
        for i, button in enumerate(self.russian_buttons):
            if i == self.current_selection:
                button.config(style="Selected.TButton")
            else:
                button.config(style="TButton")
        
        # Update English translation
        if 0 <= self.current_selection < len(self.words):
            english_word = self.words[self.current_selection][1]
            self.english_var.set(f"English: {english_word}")
        
    def move_selection(self, direction):
        """Move selection in the given direction for dynamic columns layout with auto-scroll"""
        total_words = len(self.words)
        if total_words == 0:
            return
            
        old_selection = self.current_selection
        columns = self.current_columns
        
        if direction == 'up' and self.current_selection >= columns:
            self.current_selection -= columns
        elif direction == 'down' and self.current_selection < total_words - columns:
            self.current_selection += columns
        elif direction == 'left' and self.current_selection > 0:
            self.current_selection -= 1
        elif direction == 'right' and self.current_selection < total_words - 1:
            self.current_selection += 1
        
        # Only update if selection changed
        if self.current_selection != old_selection:
            self.update_selection()
            self.auto_scroll_to_selection()

    def auto_scroll_to_selection(self):
        """Automatically scroll to make the selected item visible"""
        if not hasattr(self, 'canvas') or not hasattr(self, 'scrollable_frame'):
            return
            
        try:
            # Get the selected button
            if 0 <= self.current_selection < len(self.russian_buttons):
                button = self.russian_buttons[self.current_selection]
                
                # Get button position relative to scrollable frame
                button_y = button.winfo_y()
                button_height = button.winfo_height()
                
                # Get canvas viewport info
                canvas_height = self.canvas.winfo_height()
                current_scroll = self.canvas.yview()[0]  # Current scroll position (0-1)
                total_height = self.scrollable_frame.winfo_height()
                
                # Calculate visible range in pixels
                visible_top = current_scroll * total_height
                visible_bottom = visible_top + canvas_height
                
                # Check if button is outside visible area
                if button_y < visible_top:
                    # Button is above visible area - scroll up
                    target_scroll = button_y / total_height
                    self.canvas.yview_moveto(target_scroll)
                elif button_y + button_height > visible_bottom:
                    # Button is below visible area - scroll down
                    target_scroll = (button_y + button_height - canvas_height) / total_height
                    self.canvas.yview_moveto(target_scroll)
                    
        except Exception as e:
            print(f"Auto-scroll error: {e}")
        
    def on_level_changed(self):
        """Callback when level selection changes"""
        level = self.level_var.get()
        self.load_level_data(level)
        self.refresh_word_list()
        self.update_status(f"Switched to level {level}")
        
    def refresh_word_list(self):
        """Refresh the word list display with dynamic columns"""
        # Clear existing widgets
        for widget in self.word_frame.winfo_children():
            widget.destroy()
        
        # Use current columns
        columns = self.current_columns
        
        # Recreate Russian word buttons in dynamic columns
        self.russian_buttons = []
        for i, (russian, english) in enumerate(self.words):
            row = i // columns  # Calculate row
            col = i % columns   # Calculate column
            
            button = ttk.Button(
                self.word_frame, 
                text=russian,
                width=20,
                command=lambda idx=i: self.select_word(idx)
            )
            button.grid(row=row, column=col, padx=5, pady=2, sticky="w")
            self.russian_buttons.append(button)
        
        # Update selection
        self.current_selection = 0
        self.update_selection()
        
    def on_window_resize(self, event):
        """当窗口大小改变时自动调整列数"""
        # 只有当是主窗口改变大小时才处理
        if event.widget == self.root:
            self.adjust_columns_based_on_width()

    def adjust_columns_based_on_width(self):
        """根据窗口宽度调整列数"""
        try:
            # 获取主窗口宽度
            window_width = self.root.winfo_width()
            
            # 计算可用宽度（减去边距和滚动条宽度）
            available_width = window_width - 5  # 减去边距和滚动条的大概宽度
            
            # 计算合适的列数
            button_min_width = 120  # 每个按钮的最小宽度（像素）
            max_columns = max(1, available_width // button_min_width)
            
            # 限制最大列数不超过8，最小不少于2
            new_columns = min(18, max(2, max_columns))
            
            # 如果列数发生变化，重新布局
            if new_columns != self.current_columns:
                self.current_columns = new_columns
                self.refresh_word_list()
                
        except Exception as e:
            print(f"Adjust columns error: {e}")

    def on_mousewheel(self, event):
        """处理鼠标滚轮事件"""
        # Windows
        if event.delta:
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        # Linux
        elif event.num == 4:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.canvas.yview_scroll(1, "units")
        
    def bind_keyboard_events(self):
        """Bind keyboard events for navigation"""
        self.root.bind('h', lambda e: self.move_selection('left'))
        self.root.bind('j', lambda e: self.move_selection('down'))
        self.root.bind('k', lambda e: self.move_selection('up'))
        self.root.bind('l', lambda e: self.move_selection('right'))
        self.root.bind('<Return>', lambda e: self.play_current_word())
        self.root.focus_set()
        
    def create_widgets(self):
        """Create UI widgets"""
        # Configure styles
        style = ttk.Style()
        style.configure("Selected.TButton", background="red", foreground="blue")
        
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Title
        title_label = ttk.Label(main_frame, text="Russian Vocabulary", 
                               font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=10)
        
        # Settings frame
        settings_frame = ttk.LabelFrame(main_frame, text="Settings", padding="10")
        settings_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        # Level selection
        level_frame = ttk.Frame(settings_frame)
        level_frame.grid(row=0, column=0, columnspan=2, sticky="w", pady=5)
        
        ttk.Label(level_frame, text="Level:").grid(row=0, column=0, padx=5)
        
        # Create level radio buttons
        for i, level in enumerate(self.levels):
            rb = ttk.Radiobutton(
                level_frame, 
                text=level,
                variable=self.level_var,
                value=level,
                command=self.on_level_changed
            )
            rb.grid(row=0, column=i+1, padx=5)
        
        # Proxy settings
        proxy_frame = ttk.Frame(settings_frame)
        proxy_frame.grid(row=1, column=0, columnspan=2, sticky="w", pady=5)
        
        proxy_check = ttk.Checkbutton(proxy_frame, text="Use Proxy", 
                                     variable=self.use_proxy)
        proxy_check.grid(row=0, column=0, padx=5)
        
        proxy_entry = ttk.Entry(proxy_frame, textvariable=self.proxy_url, width=40)
        proxy_entry.grid(row=0, column=1, padx=5)
        
        # Download button
        self.download_btn = ttk.Button(settings_frame, text="Download All Audio for Current Level", 
                                      command=self.on_download_all)
        self.download_btn.grid(row=2, column=0, columnspan=2, pady=5)
        
        # Russian words frame
        russian_frame = ttk.LabelFrame(main_frame, text=f"Russian Words - Level {self.level_var.get()} - Use h/j/k/l to navigate, Enter to play", padding="10")
        russian_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # Create scrollable frame for Russian words
        russian_container = ttk.Frame(russian_frame)
        russian_container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Create canvas and scrollbar for Russian words
        self.canvas = tk.Canvas(russian_container, height=300)
        scrollbar = ttk.Scrollbar(russian_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        # 绑定鼠标滚轮事件到 Canvas
        def on_canvas_mousewheel(event):
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        
        def on_canvas_linux_scroll_up(event):
            self.canvas.yview_scroll(-1, "units")
        
        def on_canvas_linux_scroll_down(event):
            self.canvas.yview_scroll(1, "units")
        
        # Windows 鼠标滚轮
        self.canvas.bind("<MouseWheel>", on_canvas_mousewheel)
        # Linux 鼠标滚轮
        self.canvas.bind("<Button-4>", on_canvas_linux_scroll_up)
        self.canvas.bind("<Button-5>", on_canvas_linux_scroll_down)
        
        # 同时绑定到整个应用程序，确保在任何位置都能滚动
        def on_root_mousewheel(event):
            # 检查事件是否发生在单词列表区域内
            if (event.widget == self.canvas or 
                event.widget == self.scrollable_frame or 
                event.widget == self.word_frame or
                str(event.widget).startswith(str(self.canvas))):
                on_canvas_mousewheel(event)
        
        def on_root_linux_scroll_up(event):
            if (event.widget == self.canvas or 
                event.widget == self.scrollable_frame or 
                event.widget == self.word_frame or
                str(event.widget).startswith(str(self.canvas))):
                on_canvas_linux_scroll_up(event)
        
        def on_root_linux_scroll_down(event):
            if (event.widget == self.canvas or 
                event.widget == self.scrollable_frame or 
                event.widget == self.word_frame or
                str(event.widget).startswith(str(self.canvas))):
                on_canvas_linux_scroll_down(event)
        
        # 绑定到根窗口
        self.root.bind("<MouseWheel>", on_root_mousewheel)
        self.root.bind("<Button-4>", on_root_linux_scroll_up)
        self.root.bind("<Button-5>", on_root_linux_scroll_down)
        
        # 绑定窗口大小变化事件
        self.root.bind('<Configure>', self.on_window_resize)
        
        # Word frame inside scrollable frame
        self.word_frame = ttk.Frame(self.scrollable_frame)
        self.word_frame.pack(fill="both", expand=True)
        
        # Pack canvas and scrollbar
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # English translation area (single line at the bottom)
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        # English translation label
        self.english_var = tk.StringVar(value="English: ")
        english_label = ttk.Label(bottom_frame, textvariable=self.english_var, 
                                 font=("Arial", 14, "bold"))
        english_label.grid(row=0, column=0, padx=10, sticky="w")
        
        # Status bar
        status_frame = ttk.Frame(main_frame)
        status_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        status_label = ttk.Label(status_frame, textvariable=self.status_var, 
                                relief="sunken", anchor="w")
        status_label.pack(fill="x", padx=5, pady=2)
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        russian_frame.columnconfigure(0, weight=1)
        russian_frame.rowconfigure(0, weight=1)
        russian_container.columnconfigure(0, weight=1)
        russian_container.rowconfigure(0, weight=1)
        
    def select_word(self, index):
        """Select word when clicked with mouse"""
        self.current_selection = index
        self.update_selection()
        self.play_current_word()

if __name__ == "__main__":
    root = tk.Tk()
    app = RussianVocabularyApp(root)
    root.mainloop()